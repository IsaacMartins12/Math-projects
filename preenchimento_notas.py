import time
import openpyxl

#Lista que guarda todos as planilhas que contém notas de formulários

planilhas = ['F2- números inteiros.xlsx','F3- Frações.xlsx','F4- decimais.xlsx','F5- Expressões_Frações Algébricas.xlsx',
            'F6- expressoes ll.xlsx','F7- Conjuntos Numéricos.xlsx','F8- Conjuntos Numericos - Parte 2.xlsx',
            'F9-  Teoria dos Conjuntos.xlsx','F10- Teoria dos Conjuntos ll.xlsx'
            ]

pessoas = [] # Nessa lista voce adiciona as pessoas que deseja procurar as notas

notas = {}
achou = 0
nota=0
indice = 48

workbook2 = openpyxl.load_workbook('NOTAS 306_211.(2).xlsx') #Arquivo que guardará as notas achadas

worksheet2 = workbook2.active

for k in pessoas : 
    
 nome = k.lower().strip()

 for j in planilhas :
        
    # abre o arquivo Excel
    workbook = openpyxl.load_workbook(j)

    # acessa a planilha desejada
    worksheet = workbook.active

    # percorre as células na planilha
    for row in worksheet.iter_rows():
        for cell in row:
            if str(cell.value).lower() == nome or nome in str(cell.value).lower() or nome== str(cell.value).lower():
                linha = cell.coordinate[1::]
                achou = 1
                value = worksheet['C'+linha].value
                notas[nota+2] = value

    if (achou==0) : notas[nota+2] = 0
    
    achou = 0
    nota=nota+1
    
 worksheet2['A'+ str(indice)] = k
 worksheet2['B'+ str(indice)] = str(notas)
 notas = {}
 achou = 0
 nota=0
 time.sleep(0.2)
 
 indice+=1
 # Salva o arquivo do Excel

workbook2.save(filename="NOTAS 306_211.(3).xlsx")
